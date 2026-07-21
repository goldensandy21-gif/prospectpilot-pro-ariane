import csv
import io

from openpyxl import load_workbook

from ..models import Prospect, ProspectEvidence
from .enrichment import source_for, normalize_value, verify_email, verify_phone


HEADER_MAP = {
    "entreprise": "name",
    "nom": "name",
    "company": "name",
    "raison sociale": "legal_name",
    "site": "website",
    "site web": "website",
    "website": "website",
    "secteur": "sector",
    "activité": "sector",
    "naf": "naf_code",
    "code naf": "naf_code",
    "adresse": "address",
    "ville": "city",
    "pays": "country",
    "code postal": "postal_code",
    "email": "public_email",
    "e-mail": "public_email",
    "mail": "public_email",
    "telephone": "public_phone",
    "téléphone": "public_phone",
    "tel": "public_phone",
    "siren": "siren",
    "siret": "siret",
    "effectif": "employee_band",
}


def normalize_header(value):
    return str(value or "").strip().lower().replace("_", " ")


def map_row(raw):
    mapped = {}
    for key, value in raw.items():
        field = HEADER_MAP.get(normalize_header(key))
        if field and value not in (None, ""):
            mapped[field] = str(value).strip()
    return mapped


def parse_csv(uploaded_file):
    content = uploaded_file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    return [map_row(row) for row in reader]


def parse_xlsx(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x or "") for x in rows[0]]
    out = []
    for row in rows[1:]:
        out.append(map_row(dict(zip(headers, row))))
    return out


def parse_import_file(uploaded_file):
    name = (uploaded_file.name or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(uploaded_file)
    return parse_csv(uploaded_file)


def import_prospects_from_upload(uploaded_file, owner=None):
    rows = [row for row in parse_import_file(uploaded_file) if row.get("name") or row.get("siret") or row.get("website")]
    source = source_for("user_import")
    created = 0
    updated = 0
    prospects = []

    for row in rows:
        email = row.pop("public_email", "")
        phone = row.pop("public_phone", "")
        row.setdefault("name", row.get("website") or row.get("siret") or "Prospect importé")
        lookup = {"siret": row.get("siret")} if row.get("siret") else {"name": row.get("name"), "website": row.get("website", "")}
        defaults = {
            **row,
            "owner": owner,
            "source": "user_import",
            "prospecting_allowed": True,
        }
        if email:
            defaults["public_email"] = email.strip().lower()
        if phone:
            defaults["public_phone"] = phone.strip()

        prospect, was_created = Prospect.objects.update_or_create(defaults=defaults, **lookup)
        created += 1 if was_created else 0
        updated += 0 if was_created else 1
        prospects.append(prospect)

        for field_name, value in {**row, "email": email, "phone": phone}.items():
            if not value:
                continue
            value_type = "email" if field_name == "email" else "phone" if field_name == "phone" else "company"
            check = verify_email(value) if value_type == "email" else verify_phone(value) if value_type == "phone" else {"status": "unverified", "confidence": 55}
            ProspectEvidence.objects.update_or_create(
                prospect=prospect,
                field_name=field_name,
                normalized_value=normalize_value(value),
                defaults={
                    "source": source,
                    "value": value,
                    "value_type": value_type,
                    "confidence_score": check.get("confidence", 55),
                    "verification_status": check.get("status", "unverified"),
                    "source_url": "",
                    "raw_payload": {"imported": True},
                    "is_current": True,
                },
            )

    return {"created": created, "updated": updated, "prospects": prospects}
