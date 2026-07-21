from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

def prospects_csv(queryset):
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(["Entreprise","Secteur","NAF","Ville","Site","E-mail","Téléphone","Formulaire","Score technique","Score commercial","Score adéquation","Priorité","Statut"])
    for p in queryset:
        form = p.contact_forms.first()
        writer.writerow([p.name,p.sector,p.naf_code,p.city,p.website,p.public_email,p.public_phone,form.page_url if form else "",p.technical_score,p.commercial_score,p.fit_score,p.priority_score,p.get_status_display()])
    return out.getvalue().encode("utf-8-sig")

def prospects_xlsx(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    headers = ["Entreprise","Secteur","NAF","Département","Ville","Site","E-mail","Téléphone","Formulaire","Technique","Commercial","Adéquation","Priorité","Statut"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True,color="FFFFFF")
        cell.fill = PatternFill("solid",fgColor="3158FF")
    for p in queryset:
        form = p.contact_forms.first()
        ws.append([p.name,p.sector,p.naf_code,p.department,p.city,p.website,p.public_email,p.public_phone,form.page_url if form else "",p.technical_score,p.commercial_score,p.fit_score,p.priority_score,p.get_status_display()])
    for col in ws.columns:
        width = min(45,max(len(str(c.value or "")) for c in col)+2)
        ws.column_dimensions[col[0].column_letter].width = width
    out = BytesIO()
    wb.save(out)
    return out.getvalue()

def prospect_pdf(prospect, summary=None, pages=None):
    out = BytesIO()
    doc = SimpleDocTemplate(out,pagesize=A4,rightMargin=38,leftMargin=38,topMargin=38,bottomMargin=38)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Small",parent=styles["BodyText"],fontSize=8,textColor=colors.HexColor("#667083")))
    story = [
        Paragraph("ProspectPilot — Rapport d’audit",styles["Title"]),
        Paragraph(prospect.name,styles["Heading1"]),
        Paragraph(f"{prospect.sector or 'Secteur non renseigné'} · {prospect.city or 'Ville non renseignée'}",styles["BodyText"]),
        Spacer(1,14),
        Table([
            ["Score technique",str(prospect.technical_score)],
            ["Score commercial",str(prospect.commercial_score)],
            ["Adéquation cible",str(prospect.fit_score)],
            ["Priorité",str(prospect.priority_score)],
        ],colWidths=[220,100],style=TableStyle([
            ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#E7ECFF")),
            ("GRID",(0,0),(-1,-1),.5,colors.HexColor("#D8D1C4")),
            ("FONTNAME",(0,0),(-1,-1),"Helvetica"),
            ("PADDING",(0,0),(-1,-1),8),
        ])),
        Spacer(1,18),
    ]
    if summary:
        story += [
            Paragraph("Synthèse du crawl",styles["Heading2"]),
            Paragraph(f"{summary.pages_crawled} pages analysées · {summary.broken_links} liens cassés · réponse moyenne {summary.average_response_ms} ms",styles["BodyText"]),
            Spacer(1,8),
            Paragraph("Problèmes détectés",styles["Heading3"]),
        ]
        for issue in summary.issues:
            story.append(Paragraph(f"• {issue}",styles["BodyText"]))
        story.append(Spacer(1,8))
        story.append(Paragraph("Recommandations",styles["Heading3"]))
        for rec in summary.recommendations:
            story.append(Paragraph(f"• {rec}",styles["BodyText"]))
        story.append(Spacer(1,14))
    if pages:
        story.append(Paragraph("Pages analysées",styles["Heading2"]))
        data = [["URL","Statut","Titre","Réponse"]]
        for p in pages[:20]:
            data.append([p.url[:55],str(p.http_status or ""),p.title[:40],f"{p.response_ms or 0} ms"])
        table = Table(data,colWidths=[230,45,150,55],repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#161B29")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),.35,colors.HexColor("#D8D1C4")),
            ("FONTSIZE",(0,0),(-1,-1),7),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
        ]))
        story.append(table)
    story += [Spacer(1,18),Paragraph("Ce rapport est une analyse automatisée à vérifier humainement avant toute recommandation commerciale.",styles["Small"])]
    doc.build(story)
    return out.getvalue()
